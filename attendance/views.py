from django.shortcuts import render
from django.views.generic.base import TemplateView
from django.contrib.auth.models import User, Group
from rest_framework import viewsets
from .models import Hostel, Mess, Student, HostelStudent, Attendance, HostelRC, Year, HostelMessStudent
from .serializers import HostelSerializer, MessSerializer, StudentSerializer, HostelStudentSerializer, AttendanceSerializer, YearSerializer
from rest_framework.views import APIView
import datetime
from rest_framework import permissions
from rest_framework.response import Response
from django.conf import settings
from django.utils import timezone
import pytz
from django.http import JsonResponse
from django.db.models import Q
# from master_api.views import get_role
from rest_framework.pagination import PageNumberPagination
from django.db import transaction
import openpyxl
from django.contrib import messages
from django.shortcuts import redirect, reverse

class StandardResultsSetPagination(PageNumberPagination):
    page_size = 10
    page_size_query_param = 'page_size'
    max_page_size = 1000
    def get_paginated_response(self, data):
        return Response({
            # 'links': {
            #    'next': self.get_next_link(),
            #    'previous': self.get_previous_link()
            # },
            'count': self.page.paginator.count,
            'total_pages': self.page.paginator.num_pages,
            'results': data
        })
class HostelView(viewsets.ModelViewSet):

    queryset = Hostel.objects.all()
    serializer_class = HostelSerializer


    def get_queryset(self):
        queryset = Hostel.objects.all()
        
        return queryset
    def perform_create(self, serializer):
    	serializer.save(created_by=self.request.user,modified_by=self.request.user,created_at=datetime.datetime.now(pytz.timezone(settings.TIME_ZONE)),modified_at=datetime.datetime.now(pytz.timezone(settings.TIME_ZONE)))
    def perform_update(self, serializer):
    	serializer.save(modified_at=datetime.datetime.now(pytz.timezone(settings.TIME_ZONE)),modified_by=self.request.user)

class MessView(viewsets.ModelViewSet):

    queryset = Mess.objects.all()
    serializer_class = MessSerializer


    def get_queryset(self):
        queryset = Mess.objects.all()
        
        return queryset
    def perform_create(self, serializer):
    	serializer.save(created_by=self.request.user,modified_by=self.request.user,created_at=datetime.datetime.now(pytz.timezone(settings.TIME_ZONE)),modified_at=datetime.datetime.now(pytz.timezone(settings.TIME_ZONE)))
    def perform_update(self, serializer):
    	serializer.save(modified_at=datetime.datetime.now(pytz.timezone(settings.TIME_ZONE)),modified_by=self.request.user)
    	
class StudentView(viewsets.ModelViewSet):

    queryset = Student.objects.all()
    serializer_class = StudentSerializer


    def get_queryset(self):
        queryset = Student.objects.all()
        
        return queryset
    def perform_create(self, serializer):
    	serializer.save(created_by=self.request.user,modified_by=self.request.user,created_at=datetime.datetime.now(pytz.timezone(settings.TIME_ZONE)),modified_at=datetime.datetime.now(pytz.timezone(settings.TIME_ZONE)))
    def perform_update(self, serializer):
    	serializer.save(modified_at=datetime.datetime.now(pytz.timezone(settings.TIME_ZONE)),modified_by=self.request.user)

class HostelStudentView(viewsets.ModelViewSet):

    queryset = HostelStudent.objects.all()
    serializer_class = HostelStudentSerializer
    pagination_class = StandardResultsSetPagination

    def get_queryset(self):
        queryset = HostelStudent.objects.all()
        
        search = self.request.query_params.get('search', None)
        purpose = self.request.query_params.get('purpose', None)
        current_year = self.request.query_params.get('current_year', None)
        if current_year:
            print("CURRR...")
            print(current_year)
            hostel_rc = HostelRC.objects.filter(rc_id=self.request.user.id,year_id=current_year).last()
            if hostel_rc:
            	queryset = queryset.filter(hostel_id = hostel_rc.hostel)
            else:
                queryset = []
        else:
            queryset = []
        # organization = self.request.query_params.get('organization', None)
        # if organization is not None:
        #     queryset = queryset.filter(organization_id=organization)
        if search:
            queryset = queryset.filter(Q(student__student_name__icontains=search)|Q(student__register_number__icontains=search))
        if purpose=='form' and queryset:
            self.pagination_class.page_size = len(queryset)
        return queryset
    def perform_create(self, serializer):
    	serializer.save(created_by=self.request.user,modified_by=self.request.user,created_at=datetime.datetime.now(pytz.timezone(settings.TIME_ZONE)),modified_at=datetime.datetime.now(pytz.timezone(settings.TIME_ZONE)))
    def perform_update(self, serializer):
    	serializer.save(modified_at=datetime.datetime.now(pytz.timezone(settings.TIME_ZONE)),modified_by=self.request.user)

class AttendanceView(viewsets.ModelViewSet):

    queryset = Attendance.objects.all()
    serializer_class = AttendanceSerializer
    pagination_class = StandardResultsSetPagination

    def get_queryset(self):
        current_date = datetime.datetime.now(tz=timezone.utc)
        
        queryset = Attendance.objects.all()
        search = self.request.query_params.get('search', None)
        purpose = self.request.query_params.get('purpose', None)
        if search:
            queryset = queryset.filter(organization_id=search)
        queryset = queryset.filter(attendance_at__startswith=datetime.date(current_date.year,current_date.month,current_date.day))
        if search:
            queryset = queryset.filter(Q(hostel_student__student__student_name__icontains=search)|Q(hostel_student__student__register_number__icontains=search))
        if purpose=='form' and queryset:
            self.pagination_class.page_size = len(queryset)
        return queryset
    def perform_create(self, serializer):
    	serializer.save(created_by=self.request.user,modified_by=self.request.user,created_at=datetime.datetime.now(pytz.timezone(settings.TIME_ZONE)),modified_at=datetime.datetime.now(pytz.timezone(settings.TIME_ZONE)))
    def perform_update(self, serializer):
    	serializer.save(modified_at=datetime.datetime.now(pytz.timezone(settings.TIME_ZONE)),modified_by=self.request.user)
class SaveAttendance(APIView):
    def get(self,request):
        pass
    def post(self,request):
        current_date = datetime.datetime.now(tz=timezone.utc)
        data = request.data
        print(data)
        for ke in data.keys():
            attendance_obj = Attendance.objects.filter(attendance_at__startswith=datetime.date(current_date.year,current_date.month,current_date.day),hostel_student_id=ke).last()
            if not attendance_obj:
                attendance_obj = Attendance()
                attendance_obj.created_by_id = request.user.id
            attendance_obj.hostel_student_id = ke
            attendance_obj.is_present = data[ke]
            attendance_obj.modified_at = datetime.datetime.now(pytz.timezone(settings.TIME_ZONE))
            attendance_obj.modified_by_id = request.user.id
            attendance_obj.save()
        datas = {"message": "success"}
        return Response(datas)

class YearView(viewsets.ModelViewSet):

    queryset = Year.objects.all()
    serializer_class = YearSerializer
    pagination_class = StandardResultsSetPagination

    def get_queryset(self):
        queryset = Year.objects.all()
        get_current_year = self.request.query_params.get('get_current_year', None)
        if get_current_year is not None:
            queryset = queryset.filter(current_year=True)
        return queryset
    def perform_create(self, serializer):
        serializer.save(created_by=self.request.user,modified_by=self.request.user,created_at=datetime.datetime.now(pytz.timezone(settings.TIME_ZONE)),modified_at=datetime.datetime.now(pytz.timezone(settings.TIME_ZONE)))
    def perform_update(self, serializer):
        serializer.save(modified_at=datetime.datetime.now(pytz.timezone(settings.TIME_ZONE)),modified_by=self.request.user)
class CheckCurrentYear(APIView):
    def get(self, request):
        data = request.GET
        year_id = data.get('year_id')
        check_year = Year.objects.filter(current_year=True).last()
        if check_year:
            if year_id:
                if str(check_year.id) == str(year_id):
                    datas = {"already_exit": False}
                else:
                    datas = {"already_exit": True}
            else:
                datas = {"already_exit": True}
        else:
            datas = {"already_exit": False}
        return Response(datas)
class CheckYear(APIView):
    def get(self, request):
        data = request.GET
        year = data.get('year')
        message = ""
        if year:
            hostel_rc = HostelRC.objects.filter(Q(year_id=year))
            if hostel_rc:
                message = "Hostel RC"
            hostel_student = HostelStudent.objects.filter(Q(year_id=year))
            if hostel_student:
                if message:
                    message = message + ", Hostel Student"
                else:
                    message = "Hostel Student"
        is_dependent = False
        if message:
            is_dependent = True
        response_data = {"is_dependent": is_dependent, "message": message}
        return JsonResponse(response_data)


class ImportView(TemplateView):
    template_name = 'base/import.html'

    def get(self, request):

        return render(request, self.template_name, {})

    # @transaction.atomic
    def post(self, request, *args, **kwargs):
        data = request.POST
        file = request.FILES
        sheet_name = data.get('sheet_name')
        wb = openpyxl.load_workbook(file.get('import_file'), data_only=True)
        hostel = Hostel.objects.filter(name=sheet_name).last()
        current_year = Year.objects.filter(current_year=True).order_by('-from_year').last()
        if current_year:
            if not hostel:
                hostel = Hostel()
                hostel.name = sheet_name
                hostel.created_by_id = request.user.id
                hostel.modified_by_id = request.user.id
                hostel.modified_at = datetime.datetime.now(pytz.timezone(settings.TIME_ZONE))
                hostel.save()
            try:
                sheet = wb.get_sheet_by_name(sheet_name)
            except Exception as e:
                messages.error(request, e)
                return redirect('import')
            i = 0
            # with transaction.atomic():
            for rowOfCellObjects in sheet:
                student = Student()
                student_year  = ''
                student_semester = ''
                room_number = ''
                check_student  = None
                if i > 0:
                    j = 1
                    for cellObj in rowOfCellObjects:
                        cell_value = str(cellObj.value)
                        # if cellObj.value:
                        #     cell_value = str(cellObj.value)
                        #     if "." in cell_value:
                        #         cell_value = cell_value.split('.')
                        #         cell_value = cell_value[0]
                        # else:
                        #     cell_value = None
                        # print(cell_value)
                        if j == 1:
                            pass
                        elif j == 2:
                            # employee_id = cell_value
                            student.student_name = cell_value
                        elif j == 3:
                            check_student = Student.objects.filter(roll_number=str(cell_value))
                            student.roll_number = cell_value
                        elif j == 4:
                            student.email = cellObj.value
                        elif j == 5:
                            student.phone_number = cell_value
                        elif j == 6:
                            gender = ''
                            if cell_value == 'MALE':
                                gender = 'male'
                            elif  cell_value == 'FEMALE':
                                gender = 'female'
                            student.gender = gender
                        elif j == 7:
                            student.dob = datetime.datetime.strptime(str(cell_value), "%Y-%m-%d %H:%M:%S").strftime(
                            "%Y-%m-%d")
                        elif j == 8:
                            student_year = cell_value
                        elif j == 9:
                            student_semester = cell_value
                        elif j == 10:
                            student.college = cell_value
                        elif j == 11:
                            student.degree = cell_value
                        elif j == 12:
                            branch = 'others'
                            if cell_value == "Computer Science and Engineering":
                                branch = 'cse'
                            elif cell_value == "Electronics and Communication Engineering":
                                branch = 'ece'
                            elif cell_value == "Electrical and Electronics Engineering":
                                branch = 'eee'
                            elif cell_value == "Information Technology":
                                branch = 'it'
                            elif cell_value == "Mining Engineering":
                                branch = 'me'
                            student.branch = branch
                        # elif j == 13:
                        #     Hostal block
                        #     student.date_of_hire = cellObj.value
                        elif j == 14:
                            # Room Number
                            room_number = cell_value
                        elif j == 15:
                            student.address = cell_value
                        elif j == 16:
                            student.fathers_name = cell_value
                        elif j == 17:
                            if cell_value:
                                student.fathers_phone_no = cell_value
                        elif j == 18:
                            student.mothers_name = cell_value
                        elif j == 19:
                            if cell_value:
                                student.mothers_phone_no = cell_value
                        # elif j == 20:
                        #     contractor.emergency_contact_person = cell_value
                        # elif j == 21:
                        #     contractor.emergency_contact_number = cell_value
                        # elif j == 22:
                        #     company = Company.objects.filter(name=cell_value).last()
                        #     if request.user.role == 'supervisor':
                        #         user_company = UserCompany.objects.filter(user=request.user.id).last()
                        #         if not user_company:
                        #             is_company_error = False
                        #         if company and user_company:
                        #             if company.id == user_company.company_id:
                        #                 contractor_company.company_id = company.id
                        #             else:
                        #                 is_company_error = False
                        #     else:
                        #         if company:
                        #             contractor_company.company_id = company.id
                        # elif j == 23:
                        #     if cellObj.value:
                        #         location_name = cellObj.value
                        #         location_name = location_name.strip()



                        
                        j += 1
                    try:
                        if not check_student:
                            student.created_by_id = request.user.id
                            student.modified_by_id = request.user.id
                            student.modified_at = datetime.datetime.now(pytz.timezone(settings.TIME_ZONE))
                            student.save()
                            # hostel_student = HostelStudent()
                            # hostel_student.year_id = current_year
                            # hostel_student.hostel_id = hostel.id
                            # hostel_student.student_id = student.id
                            # hostel_student.student_year = int(student_year)
                            # hostel_student.student_semester = student_semester
                            # hostel_student.room_number = room_number
                            # hostel_student.created_by_id = request.user.id
                            # hostel_student.modified_by_id = request.user.id
                            # hostel_student.modified_at = datetime.datetime.now(pytz.timezone(settings.TIME_ZONE))
                            # hostel_student.save()
                    except Exception as e:
                        print("Error ...")
                        print(e)
                    # try:
                    #     if employee_id:
                    #         contractor.save()
                    #         if contractor.bio_start_location_id:
                    #             check_user = CreateUsersIntoDevice.objects.filter(contractor_id=contractor.id,location_id=contractor.bio_start_location_id,process_date=datetime.datetime.now().strftime("%Y-%m-%d")).last()
                    #             if not check_user:
                    #                 create_user = CreateUsersIntoDevice()
                    #                 create_user.contractor_id = contractor.id
                    #                 create_user.location_id = contractor.bio_start_location_id
                    #                 create_user.process_date = datetime.datetime.now().strftime("%Y-%m-%d")
                    #                 create_user.save()
                    #             if not contractor.bio_start_location_id in is_immadiate_locations and is_immadiate_push:
                    #                 is_immadiate_locations.append(contractor.bio_start_location_id)
                    #             if not contractor.id in is_immadiate_employees and is_immadiate_push:
                    #                 is_immadiate_employees.append(contractor.id)
                                
                    #         contractor_company.contractor_id = contractor.id
                    #         contractor_company.start_date = contractor.date_of_hire
                    #         contractor_company.save()
                    #         for skill_id in skill_ids:
                    #             contractor_skill = ContractorSkill()
                    #             contractor_skill.skill_id = skill_id
                    #             contractor_skill.contractor_id = contractor.id
                    #             contractor_skill.save()

                    #         if project_name and location_name and department_name:
                    #             project = Project.objects.filter(name=project_name.strip(), location__name=location_name,
                    #                                              department__name=department_name,
                    #                                              place_of_work__name=place_of_work_name).last()
                    #             if project:
                    #                 if project.location_id in locations:
                    #                     pass_creation.project_id = project.id
                    #                 else:
                    #                     is_assigned_location = False
                    #         pass_creation.from_date = contractor.date_of_hire
                    #         pass_creation.contractor_id = contractor.id
                    #         pass_creation.employee_id = contractor.employee_id
                    #         pass_creation.record_type = "import"
                    #         pass_creation.save()

                    #     # print("__________")
                    #     # print(rowOfCellObjects[22].value)location
                    # except Exception as e:
                    #     error = ast.literal_eval(str(e))
                    #     if len(error) > 1:
                    #         error_message = error[1]
                    #     else:
                    #         error_message = error[0]
                    #     if not is_company_error:
                    #         error_message = "You dont have access for this company."
                    #     error_message = error_message.replace('for key', 'in the ')
                    #     error_message = error_message.replace('pan_number', 'PAN')
                    #     error_message = error_message.replace('employee_id', 'Employee ID')
                    #     error_message = error_message.replace('aadhar_number', 'Aadhaar Number')
                    #     error_message = error_message.replace('company_id', 'Company name')
                    #     error_message = error_message.replace('cannot be null', 'does not exist')
                    #     error_message = error_message.replace('project_id', 'Project')
                    #     error_message = error_message + " row number " + str(i)
                    #     messages.error(request, error_message)
                    #     return redirect('contractor_import')
                    #     # raise Exception(error[1]+" rowwww"+str(i))
                i += 1
        messages.success(request, 'Student imported successfully.')
        return redirect('import')

class MessStudentImportView(TemplateView):
    template_name = 'base/mess-student-import.html'

    def get(self, request):

        return render(request, self.template_name, {})

    # @transaction.atomic
    def post(self, request, *args, **kwargs):
        data = request.POST
        file = request.FILES
        sheet_name = data.get('sheet_name')
        wb = openpyxl.load_workbook(file.get('import_file'), data_only=True)
        mess = Mess.objects.filter(name=sheet_name).last()
        current_year = Year.objects.filter(current_year=True).order_by('-from_year').last()
        if current_year and mess:
            try:
                sheet = wb.get_sheet_by_name(sheet_name)
            except Exception as e:
                messages.error(request, e)
                return redirect('mess_student_import')
            i = 0
            # with transaction.atomic():
            for rowOfCellObjects in sheet:
                hostel_mess_student = HostelMessStudent()
                check_student  = None
                if i > 0:
                    j = 1
                    for cellObj in rowOfCellObjects:
                        cell_value = str(cellObj.value)
                        if j == 1:
                            hostel_mess_student.mess_id = mess.id
                        elif j == 2:
                            print(cell_value)
                            hostel = Hostel.objects.filter(name=cell_value).last()
                            hostel_mess_student.hostel_id = hostel.id
                        elif j == 3:
                            if cell_value:
                                cell_value = str(cell_value)
                                if "." in cell_value:
                                    cell_value = cell_value.split('.')
                                    cell_value = cell_value[0]
                            else:
                                cell_value = None
                            print(cell_value)

                            check_student = Student.objects.filter(roll_number=str(cell_value)).last()
                            hostel_mess_student.student_id = check_student.id
                        j += 1
                    try:
                        hostel_mess_student.year_id = current_year.id
                        hostel_mess_student.created_by_id = request.user.id
                        hostel_mess_student.modified_by_id = request.user.id
                        hostel_mess_student.modified_at = datetime.datetime.now(pytz.timezone(settings.TIME_ZONE))
                        hostel_mess_student.save()
                    except Exception as e:
                        print("Error ...")
                        print(e)
                i += 1
        messages.success(request, 'Mess Student imported successfully.')
        return redirect('mess_student_import')

class MessImportView(TemplateView):
    template_name = 'base/mess-import.html'

    def get(self, request):

        return render(request, self.template_name, {})

    # @transaction.atomic
    def post(self, request, *args, **kwargs):
        data = request.POST
        file = request.FILES
        sheet_name = data.get('sheet_name')
        wb = openpyxl.load_workbook(file.get('import_file'), data_only=True)
        # mess = Mess.objects.filter(name=sheet_name).last()
        current_year = Year.objects.filter(current_year=True).order_by('-from_year').last()
        if current_year:
            try:
                sheet = wb.get_sheet_by_name(sheet_name)
            except Exception as e:
                messages.error(request, e)
                return redirect('mess_import')
            i = 0
            # with transaction.atomic():
            for rowOfCellObjects in sheet:
                mess = Mess()
                mess_check  = None
                if i > 0:
                    j = 1
                    for cellObj in rowOfCellObjects:
                        cell_value = str(cellObj.value)
                        if j == 1:
                            mess_check = Mess.objects.filter(name=cell_value).last()
                            mess.name = cell_value
                        elif j == 2:
                            if cell_value == 'Veg':
                                mess.food_type = 'veg'
                            elif cell_value == 'Non Veg':
                                mess.food_type = 'non_veg'
                        j += 1
                    try:
                        mess.year_id = current_year.id
                        mess.created_by_id = request.user.id
                        mess.modified_by_id = request.user.id
                        mess.modified_at = datetime.datetime.now(pytz.timezone(settings.TIME_ZONE))
                        if not mess_check:
                            mess.save()
                    except Exception as e:
                        print("Error ...")
                        print(e)
                i += 1
        messages.success(request, 'Mess imported successfully.')
        return redirect('mess_import')


class RCImportView(TemplateView):
    template_name = 'base/rc.html'

    def get(self, request):

        return render(request, self.template_name, {})

    # @transaction.atomic
    def post(self, request, *args, **kwargs):
        data = request.POST
        file = request.FILES
        sheet_name = data.get('sheet_name')
        wb = openpyxl.load_workbook(file.get('import_file'), data_only=True)
        # mess = Mess.objects.filter(name=sheet_name).last()
        current_year = Year.objects.filter(current_year=True).order_by('-from_year').last()
        if current_year:
            try:
                sheet = wb.get_sheet_by_name(sheet_name)
            except Exception as e:
                messages.error(request, e)
                return redirect('rc_import')
            i = 0
            # with transaction.atomic():
            for rowOfCellObjects in sheet:
                print("IN ....")
                hostelrc = HostelRC()
                mess_check  = None
                check_hostel = None
                if i > 0:
                    j = 1
                    for cellObj in rowOfCellObjects:
                        cell_value = str(cellObj.value)
                        if j == 1:
                            check_hostel = Hostel.objects.filter(name=cell_value).last()
                            if check_hostel:
                                hostelrc.hostel_id = check_hostel.id
                        elif j == 2:
                            check_user = User.objects.filter(username=cell_value).last()
                            if check_user:
                                hostelrc.rc_id = check_user.id
                            
                        j += 1
                    try:
                        if check_hostel and check_user:
                            print("596 ...")
                            check_rc = HostelRC.objects.filter(hostel_id=check_hostel.id,rc_id=check_user.id,year_id = current_year.id )
                            hostelrc.year_id = current_year.id
                            hostelrc.created_by_id = request.user.id
                            hostelrc.modified_by_id = request.user.id
                            hostelrc.modified_at = datetime.datetime.now(pytz.timezone(settings.TIME_ZONE))
                            hostelrc.active = True
                            if not check_rc:
                                hostelrc.save()
                    except Exception as e:
                        print("Error ...")
                        print(e)
                i += 1
        messages.success(request, 'RC imported successfully.')
        return redirect('rc_import')